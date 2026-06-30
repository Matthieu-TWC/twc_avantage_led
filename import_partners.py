#!/usr/bin/env python3
"""
Import de partenaires depuis Excel vers Odoo via XML-RPC.

Usage:
    export ODOO_URL=https://your-odoo.com
    export ODOO_DB=your-db
    export ODOO_LOGIN=you@example.com
    export ODOO_API_KEY=your_api_key

    python3 import_partners.py Base_clients_import_Odoo_4.xlsx

Colonnes attendues (ligne 1 = en-têtes) :
    id, name, is_company, function, email, mobile, phone,
    street, street2, city, zip, country_id, category_id, vat,
    parent_id/id, user_id, notes, property_payment_term_id,
    property_product_pricelist
"""

import os
import sys
import xmlrpc.client
from collections import defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl manquant. Installez-le : pip3 install openpyxl")

URL     = os.environ.get('ODOO_URL', '').rstrip('/')
DB      = os.environ.get('ODOO_DB', '')
LOGIN   = os.environ.get('ODOO_LOGIN', '')
API_KEY = os.environ.get('ODOO_API_KEY', '')


def main():
    # ── Vérifications ─────────────────────────────────────────────────────
    missing = [k for k, v in {'ODOO_URL': URL, 'ODOO_DB': DB,
                               'ODOO_LOGIN': LOGIN, 'ODOO_API_KEY': API_KEY}.items() if not v]
    if missing:
        sys.exit(f"Variables d'environnement manquantes : {', '.join(missing)}")

    if len(sys.argv) < 2:
        sys.exit("Usage : python3 import_partners.py chemin/vers/fichier.xlsx")

    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        sys.exit(f"Fichier introuvable : {excel_path}")

    # ── Connexion XML-RPC ─────────────────────────────────────────────────
    print(f"→ Connexion à {URL} ...")
    common = xmlrpc.client.ServerProxy(f'{URL}/xmlrpc/2/common')
    uid = common.authenticate(DB, LOGIN, API_KEY, {})
    if not uid:
        sys.exit("Authentification échouée. Vérifiez URL, DB, login et clé API.")
    print("  Connecté ✔")

    models = xmlrpc.client.ServerProxy(f'{URL}/xmlrpc/2/object')

    def rpc(model, method, args=None, kw=None):
        return models.execute_kw(DB, uid, API_KEY, model, method, args or [], kw or {})

    def str_cell(v):
        return str(v).strip() if v not in (None, '') else ''

    def bool_cell(v):
        if isinstance(v, bool):
            return v
        return str(v).strip().lower() in ('true', '1', 'oui', 'yes')

    # ── Chargement des référentiels ───────────────────────────────────────
    print("→ Chargement des référentiels ...")

    country_data = rpc('res.country', 'search_read', [[]], {'fields': ['id', 'name']})
    country_by_name = {c['name'].lower(): c['id'] for c in country_data}
    # French aliases not covered by the Odoo instance language
    FR_COUNTRY_ALIASES = {
        'france': 'france', 'belgique': 'belgium', 'suisse': 'switzerland',
        'espagne': 'spain', 'maroc': 'morocco', 'tunisie': 'tunisia',
        'sénégal': 'senegal', "côte d'ivoire": "côte d'ivoire",
        'allemagne': 'germany', 'italie': 'italy', 'pays-bas': 'netherlands',
        'royaume-uni': 'united kingdom', 'états-unis': 'united states',
        'portugal': 'portugal', 'luxembourg': 'luxembourg',
    }
    for fr, en in FR_COUNTRY_ALIASES.items():
        if fr not in country_by_name and en in country_by_name:
            country_by_name[fr] = country_by_name[en]

    cat_data = rpc('res.partner.category', 'search_read', [[]], {'fields': ['id', 'name']})
    cat_by_name = {c['name']: c['id'] for c in cat_data}

    def get_or_create_category(name):
        if not name:
            return None
        if name not in cat_by_name:
            cat_by_name[name] = rpc('res.partner.category', 'create', [{'name': name}])
        return cat_by_name[name]

    pt_data = rpc('account.payment.term', 'search_read', [[]], {'fields': ['id', 'name']})
    pt_by_name = {p['name']: p['id'] for p in pt_data}

    def resolve_payment_term(name):
        if not name:
            return None
        if name in pt_by_name:
            return pt_by_name[name]
        return next((i for n, i in pt_by_name.items()
                     if name.lower() in n.lower() or n.lower() in name.lower()), None)

    pl_data = rpc('product.pricelist', 'search_read', [[]], {'fields': ['id', 'name']})
    pl_by_name = {p['name']: p['id'] for p in pl_data}

    def resolve_pricelist(name):
        if not name:
            return None
        if name in pl_by_name:
            return pl_by_name[name]
        return next((i for n, i in pl_by_name.items()
                     if name.lower() in n.lower() or n.lower() in name.lower()), None)

    user_data = rpc('res.users', 'search_read', [[]], {'fields': ['id', 'login', 'name']})
    user_by_login = {u['login']: u['id'] for u in user_data}
    user_by_name  = {u['name']:  u['id'] for u in user_data}

    USER_ALIASES = {
        'Thomas-Delcourt':      'Benjamin ORIOL',
        'cyrille-avantage-led': 'Cyrille FOLIE-DESJARDINS',
    }

    def resolve_user(login_str):
        if not login_str:
            return None
        resolved = USER_ALIASES.get(login_str, login_str)
        return user_by_login.get(resolved) or user_by_name.get(resolved)

    # ── Champs disponibles sur res.partner ───────────────────────────────
    partner_fields = set(rpc('res.partner', 'fields_get', [], {'attributes': ['string']}).keys())
    expected_fields = {'function', 'email', 'phone', 'street', 'street2',
                       'city', 'zip', 'vat', 'comment', 'country_id', 'category_id',
                       'user_id', 'property_payment_term_id', 'property_product_pricelist'}
    missing_fields = expected_fields - partner_fields
    if missing_fields:
        print(f"  ⚠ Champs absents sur cette instance (données ignorées) : {', '.join(sorted(missing_fields))}")

    # ── Partenaires existants ─────────────────────────────────────────────
    print("→ Chargement des partenaires existants ...")
    existing = rpc('res.partner', 'search_read', [[]], {'fields': ['id', 'name']})
    partner_by_name = {p['name']: p['id'] for p in existing}

    # ── Lecture du fichier Excel ──────────────────────────────────────────
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb['Sheet']
    all_rows = [r for r in sheet.iter_rows(min_row=2, values_only=True) if any(r)]
    total = len(all_rows)
    print(f"  {total} lignes à traiter")

    # Columns (0-based):
    # 0:id  1:name  2:is_company  3:function  4:email  5:mobile  6:phone
    # 7:street  8:street2  9:city  10:zip  11:country_id  12:category_id
    # 13:vat  14:parent_id/id  15:user_id  16:notes
    # 17:property_payment_term_id  18:property_product_pricelist

    # Tracks unresolved reference values: {column: {bad_value: [record_name, ...]}}
    # Only populated for records that are actually created (not for already-existing ones).
    unresolved = {
        'country_id':                  defaultdict(list),
        'user_id':                     defaultdict(list),
        'property_product_pricelist':  defaultdict(list),
        'property_payment_term_id':    defaultdict(list),
    }

    def build_vals(row, record_name, parent_odoo_id=None):
        r = (row + (None,) * 19)[:19]
        (_, name, is_company, function, email, mobile, phone,
         street, street2, city, zip_code, country, category, vat,
         _parent_ext, user_login, notes, payment_term, pricelist) = r

        vals = {
            'name':       str_cell(name),
            'is_company': bool_cell(is_company),
        }

        phone_val = str_cell(phone) or str_cell(mobile)
        for field, val in [
            ('function', str_cell(function)),
            ('email',    str_cell(email)),
            ('phone',    phone_val),
            ('street',   str_cell(street)),
            ('street2',  str_cell(street2)),
            ('city',     str_cell(city)),
            ('zip',      str_cell(zip_code)),
            ('vat',      str_cell(vat)),
            ('comment',  str_cell(notes)),
        ]:
            if val and field in partner_fields:
                vals[field] = val

        field_warns = []

        country_str = str_cell(country)
        cid = country_by_name.get(country_str.lower())
        if cid and 'country_id' in partner_fields:
            vals['country_id'] = cid
        elif country_str:
            unresolved['country_id'][country_str].append(record_name)
            field_warns.append(f'country_id="{country_str}"')

        cat_id = get_or_create_category(str_cell(category))
        if cat_id and 'category_id' in partner_fields:
            vals['category_id'] = [(6, 0, [cat_id])]

        user_str = str_cell(user_login)
        user_id = resolve_user(user_str)
        if user_id and 'user_id' in partner_fields:
            vals['user_id'] = user_id
        elif user_str and 'user_id' not in partner_fields:
            pass  # field missing on this instance, silently skip
        elif user_str:
            unresolved['user_id'][user_str].append(record_name)
            field_warns.append(f'user_id="{user_str}"')

        pt_str = str_cell(payment_term)
        pt_id = resolve_payment_term(pt_str)
        if pt_id and 'property_payment_term_id' in partner_fields:
            vals['property_payment_term_id'] = pt_id
        elif pt_str and 'property_payment_term_id' in partner_fields:
            unresolved['property_payment_term_id'][pt_str].append(record_name)
            field_warns.append(f'property_payment_term_id="{pt_str}"')

        pl_str = str_cell(pricelist)
        pl_id = resolve_pricelist(pl_str)
        if pl_id and 'property_product_pricelist' in partner_fields:
            vals['property_product_pricelist'] = pl_id
        elif pl_str and 'property_product_pricelist' in partner_fields:
            unresolved['property_product_pricelist'][pl_str].append(record_name)
            field_warns.append(f'property_product_pricelist="{pl_str}"')

        if parent_odoo_id:
            vals['parent_id'] = parent_odoo_id

        return vals, field_warns

    ignored_records = []  # list of (name, reason) for final report

    def process_row(row, idx, total, parent_odoo_id=None):
        name   = str_cell(row[1])
        ext_id = str_cell(row[0])
        is_co  = bool_cell(row[2])
        kind   = 'entreprise' if is_co else 'contact'
        prefix = f'[{idx:4d}/{total}]'

        # Existing check before build_vals so unresolved tracking is never
        # inflated by records that are skipped.
        if name in partner_by_name:
            odoo_id = partner_by_name[name]
            if ext_id:
                ext_id_map[ext_id] = odoo_id
            # Fill in missing fields on the existing record
            country_str = str_cell(row[11])
            cid = country_by_name.get(country_str.lower()) if country_str else None
            user_str = str_cell(row[15])
            uid_ = resolve_user(user_str) if user_str else None

            updates = {}
            update_log = []
            if cid or uid_:
                current = rpc('res.partner', 'read', [[odoo_id]], {'fields': ['country_id', 'user_id']})
                if current:
                    if cid and not current[0]['country_id']:
                        updates['country_id'] = cid
                        update_log.append(f'country_id="{country_str}"')
                    if uid_ and not current[0]['user_id']:
                        updates['user_id'] = uid_
                        update_log.append(f'user_id="{user_str}"')
            if updates:
                rpc('res.partner', 'write', [[odoo_id], updates])
                print(f'{prefix} ≡ EXISTANT  {name} ({kind})  ✏ mis à jour : {", ".join(update_log)}')
            else:
                print(f'{prefix} ≡ EXISTANT  {name} ({kind})')
            return 'existing'

        vals, field_warns = build_vals(row, name, parent_odoo_id)
        try:
            odoo_id = rpc('res.partner', 'create', [vals])
        except Exception as e:
            print(f'{prefix} ✗ ERREUR    {name} ({kind})  → {e}')
            ignored_records.append((name, str(e)))
            return 'error'

        partner_by_name[name] = odoo_id
        if ext_id:
            ext_id_map[ext_id] = odoo_id

        warn_suffix = ('  ⚠ ' + ', '.join(field_warns)) if field_warns else ''
        print(f'{prefix} ✔ CRÉÉ      {name} ({kind}){warn_suffix}')
        return 'created'

    # ── Phase 1 : enregistrements sans parent ────────────────────────────
    print()
    print("── Phase 1 : enregistrements sans parent ─────────────────────")
    ext_id_map = {}
    p1_created = p1_existing = p1_errors = 0
    p1_rows = [(i + 1, row) for i, row in enumerate(all_rows) if not str_cell(row[14])]

    for idx, row in p1_rows:
        name = str_cell(row[1])
        if not name:
            continue
        result = process_row(row, idx, total)
        if result == 'created':
            p1_created += 1
        elif result == 'error':
            p1_errors += 1
        else:
            p1_existing += 1

    # ── Phase 2 : enregistrements avec parent ────────────────────────────
    print()
    print("── Phase 2 : enregistrements avec parent ─────────────────────")
    p2_created = p2_existing = p2_skipped = p2_errors = 0
    p2_rows = [(i + 1, row) for i, row in enumerate(all_rows) if str_cell(row[14])]

    for idx, row in p2_rows:
        name          = str_cell(row[1])
        parent_ext_id = str_cell(row[14])
        if not name:
            continue

        parent_odoo_id = ext_id_map.get(parent_ext_id)
        if parent_odoo_id is None:
            print(f'[{idx:4d}/{total}] ✗ IGNORÉ    {name} → parent "{parent_ext_id}" introuvable')
            ignored_records.append((name, f'parent "{parent_ext_id}" introuvable'))
            p2_skipped += 1
            continue

        result = process_row(row, idx, total, parent_odoo_id)
        if result == 'created':
            p2_created += 1
        elif result == 'error':
            p2_errors += 1
        else:
            p2_existing += 1

    # ── Rapport final ─────────────────────────────────────────────────────
    total_created   = p1_created  + p2_created
    total_existing  = p1_existing + p2_existing
    total_errors    = p1_errors   + p2_errors
    total_skipped   = p2_skipped
    total_processed = total_created + total_existing + total_errors + total_skipped

    print()
    print('═' * 55)
    print('  RAPPORT FINAL')
    print('═' * 55)
    print(f'  Enregistrements dans le fichier  : {total:4d}')
    print(f'  Enregistrements traités          : {total_processed:4d}')
    print(f'    dont importés (créés)          : {total_created:4d}')
    print(f'    dont déjà existants (ignorés)  : {total_existing:4d}')
    print(f'    dont en erreur                 : {total_errors:4d}')
    print(f'    dont ignorés (parent manquant) : {total_skipped:4d}')

    any_unresolved = any(v for v in unresolved.values())
    if any_unresolved:
        print()
        print('  Référentiels non résolus (champs ignorés à l\'import) :')
        for column, records_by_val in unresolved.items():
            if not records_by_val:
                continue
            for val, names in sorted(records_by_val.items(), key=lambda x: -len(x[1])):
                n = len(names)
                print(f'    ⚠ {column} = "{val}"  ({n} enregistrement{"s" if n > 1 else ""} créé{"s" if n > 1 else ""}) :')
                for rec_name in names[:10]:
                    print(f'        - {rec_name}')
                if n > 10:
                    print(f'        ... et {n - 10} autres')
    else:
        print()
        print('  Référentiels non résolus : aucun ✔')

    if ignored_records:
        print()
        print(f'  Enregistrements ignorés ({len(ignored_records)}) :')
        for name, reason in ignored_records:
            print(f'    ✗ "{name}" → {reason}')

    print('═' * 55)


if __name__ == '__main__':
    main()
