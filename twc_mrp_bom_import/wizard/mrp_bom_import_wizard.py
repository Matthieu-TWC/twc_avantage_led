import base64
import io
import logging

from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class MrpBomImportWizard(models.TransientModel):
    _name = 'twc.mrp.bom.import.wizard'
    _description = 'Import de nomenclatures depuis Excel'

    def _auto_init(self):
        result = super()._auto_init()
        self._ensure_access_rights()
        return result

    @api.model
    def _ensure_access_rights(self):
        IrModelAccess = self.env['ir.model.access']
        model = self.env['ir.model']._get(self._name)
        if not model or IrModelAccess.search([('model_id', '=', model.id)], limit=1):
            return
        group = self.env.ref('mrp.group_mrp_manager', raise_if_not_found=False)
        IrModelAccess.create({
            'name': self._name,
            'model_id': model.id,
            'group_id': group.id if group else False,
            'perm_read': True,
            'perm_write': True,
            'perm_create': True,
            'perm_unlink': True,
        })

    file_data = fields.Binary(string='Fichier Excel (.xlsx)', required=True, attachment=False)
    file_name = fields.Char(string='Nom du fichier')
    state = fields.Selection(
        [('draft', 'Prêt'), ('done', 'Terminé')],
        default='draft',
    )
    result_log = fields.Text(string="Rapport d'import", readonly=True)

    def action_import(self):
        self.ensure_one()
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        raw = base64.b64decode(self.file_data)
        wb = load_workbook(filename=io.BytesIO(raw), data_only=True)

        log_lines = []

        partner_map, partner_log = self._import_partners(wb)
        log_lines.extend(partner_log)

        product_map, prod_log = self._import_finished_products(wb)
        log_lines.extend(prod_log)

        comp_map, comp_log = self._import_components(wb, partner_map)
        log_lines.extend(comp_log)

        bom_log = self._import_boms(wb, product_map, comp_map)
        log_lines.extend(bom_log)

        self.write({
            'state': 'done',
            'result_log': '\n'.join(log_lines),
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    # ------------------------------------------------------------------
    # Phase 0 — Partners
    # ------------------------------------------------------------------

    def _import_partners(self, wb):
        sheet = wb['Partenaires']
        log = []
        partner_map = {}
        created = 0
        existing = 0

        all_partners = self.env['res.partner'].search([])
        by_name = {p.name: p for p in all_partners}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            nom, est_entreprise, rang_fournisseur = (row + (None,) * 3)[:3]
            nom = (nom or '').strip()
            if not nom:
                continue

            partner = by_name.get(nom)
            if partner:
                existing += 1
            else:
                partner = self.env['res.partner'].create({
                    'name': nom,
                    'is_company': bool(est_entreprise),
                })
                by_name[nom] = partner
                created += 1

            partner_map[nom] = partner

        log.append(f'✔ Partenaires importés : {created} créés, {existing} déjà existants')
        return partner_map, log

    # ------------------------------------------------------------------
    # Phase 1 — Finished products
    # ------------------------------------------------------------------

    def _import_finished_products(self, wb):
        sheet = wb['Produits finis']
        log = []
        product_map = {}
        created = 0
        existing = 0

        all_templates = self.env['product.template'].search([])
        by_code = {t.default_code: t for t in all_templates if t.default_code}
        by_name = {t.name: t for t in all_templates}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            nom, reference_interne, uom_str, product_type = (row + (None,) * 4)[:4]
            nom = (nom or '').strip()
            if not nom:
                continue
            reference_interne = (
                str(reference_interne).strip() if reference_interne not in (None, '') else ''
            )

            tmpl = None
            if reference_interne:
                tmpl = by_code.get(reference_interne)
            if tmpl is None:
                tmpl = by_name.get(nom)

            if tmpl:
                existing += 1
            else:
                vals = {
                    'name': nom,
                    'type': self._map_product_type(product_type),
                    'uom_id': self._resolve_uom(uom_str).id,
                }
                if reference_interne:
                    vals['default_code'] = reference_interne
                tmpl = self.env['product.template'].create(vals)
                by_code[reference_interne] = tmpl
                by_name[nom] = tmpl
                created += 1

            display = f'[{reference_interne}] {nom}' if reference_interne else nom
            product_map[display] = tmpl
            product_map[nom] = tmpl

        log.append(f'✔ Produits finis       : {created} créés, {existing} déjà existants')
        return product_map, log

    # ------------------------------------------------------------------
    # Phase 2 — Components (pre-import before BOMs)
    # ------------------------------------------------------------------

    def _import_components(self, wb, partner_map):
        """Scan Nomenclatures sheet, create all missing component products.
        Returns comp_map: {(ref or '') + '|' + nom -> product.product variant}
        All creates happen here, outside any BOM savepoint, so variants are
        committed and reliably findable when the BOM phase runs.
        """
        sheet = wb['Nomenclatures']
        log = []
        created = 0
        existing = 0

        # Exact-match lookups by internal ref and by name
        all_templates = self.env['product.template'].search([])
        by_code = {t.default_code: t for t in all_templates if t.default_code}
        by_name = {t.name: t for t in all_templates}

        comp_map = {}  # key -> product.product variant

        def _key(ref, nom):
            return f'{ref}|{nom}' if ref else nom

        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not any(row):
                continue
            (
                _bom_pf, _bom_ref,
                composant_nom, composant_reference, composant_uom,
                _composant_quantite, composant_prix, composant_poids,
                composant_fournisseur, _notes
            ) = (row + (None,) * 10)[:10]

            nom = (str(composant_nom).strip() if composant_nom not in (None, '') else '')
            if not nom:
                continue
            ref = (str(composant_reference).strip() if composant_reference not in (None, '') else '')
            key = _key(ref, nom)

            if key in comp_map:
                continue  # already processed this component

            # Strict lookup: ref first, then exact name
            tmpl = None
            if ref:
                tmpl = by_code.get(ref)
            if tmpl is None:
                tmpl = by_name.get(nom)

            if tmpl:
                existing += 1
            else:
                try:
                    prix_val = float(composant_prix) if composant_prix not in (None, '') else None
                except (ValueError, TypeError):
                    prix_val = None
                try:
                    poids_val = float(composant_poids) if composant_poids not in (None, '') else None
                except (ValueError, TypeError):
                    poids_val = None

                vals = {
                    'name': nom,
                    'type': 'consu',
                    'uom_id': self._resolve_uom(composant_uom).id,
                }
                if ref:
                    vals['default_code'] = ref
                if prix_val is not None:
                    vals['standard_price'] = prix_val
                if poids_val is not None:
                    vals['weight'] = poids_val

                tmpl = self.env['product.template'].create(vals)
                by_name[nom] = tmpl
                if ref:
                    by_code[ref] = tmpl
                created += 1

            # Get variant — created automatically by Odoo alongside the template
            variant = self.env['product.product'].search(
                [('product_tmpl_id', '=', tmpl.id)], limit=1
            )
            if not variant:
                _logger.warning('No variant found for component template %s (id=%s)', nom, tmpl.id)
                continue

            comp_map[key] = variant

            # Link supplier
            fournisseur = (str(composant_fournisseur).strip()
                           if composant_fournisseur not in (None, '') else '')
            if fournisseur:
                partner = partner_map.get(fournisseur)
                if partner and not any(s.partner_id.id == partner.id for s in tmpl.seller_ids):
                    try:
                        prix_val = float(composant_prix) if composant_prix not in (None, '') else None
                    except (ValueError, TypeError):
                        prix_val = None
                    uom = self._resolve_uom(composant_uom)
                    tmpl.write({'seller_ids': [(0, 0, {
                        'partner_id': partner.id,
                        'price': prix_val or 0.0,
                        'product_uom_id': uom.id,
                    })]})

        log.append(f'✔ Composants importés  : {created} créés, {existing} déjà existants')
        return comp_map, log

    # ------------------------------------------------------------------
    # Phase 3 — BOMs
    # ------------------------------------------------------------------

    def _import_boms(self, wb, product_map, comp_map):
        sheet = wb['Nomenclatures']
        log = []
        warnings = []
        created = 0
        skipped = 0

        def _key(ref, nom):
            return f'{ref}|{nom}' if ref else nom

        current_product = None
        current_code = None
        current_lines = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not any(row):
                continue
            (
                bom_produit_fini, bom_reference,
                composant_nom, composant_reference, composant_uom,
                composant_quantite, _composant_prix, _composant_poids,
                _composant_fournisseur, _notes
            ) = (row + (None,) * 10)[:10]

            bom_produit_fini = (
                str(bom_produit_fini).strip() if bom_produit_fini not in (None, '') else ''
            )
            composant_nom = (
                str(composant_nom).strip() if composant_nom not in (None, '') else ''
            )
            composant_reference = (
                str(composant_reference).strip()
                if composant_reference not in (None, '') else ''
            )

            if bom_produit_fini:
                if current_product is not None:
                    ok, warn = self._save_bom(
                        current_product, current_code, current_lines,
                        product_map, comp_map, _key,
                    )
                    if ok == 'created':
                        created += 1
                    elif ok == 'skipped':
                        skipped += 1
                    warnings.extend(warn)
                current_product = bom_produit_fini
                current_code = (
                    str(bom_reference).strip() if bom_reference not in (None, '') else ''
                )
                current_lines = []

            if composant_nom:
                try:
                    qty_val = float(composant_quantite) if composant_quantite not in (None, '') else 1.0
                except (ValueError, TypeError):
                    qty_val = 1.0
                current_lines.append({
                    'nom': composant_nom,
                    'reference': composant_reference,
                    'qty': qty_val,
                    'row': row_idx,
                })

        if current_product is not None:
            ok, warn = self._save_bom(
                current_product, current_code, current_lines,
                product_map, comp_map, _key,
            )
            if ok == 'created':
                created += 1
            elif ok == 'skipped':
                skipped += 1
            warnings.extend(warn)

        log.append(f'✔ BOMs importées       : {created} créées, {skipped} déjà existantes')
        if warnings:
            log.append(f'⚠ Avertissements ({len(warnings)}) :')
            log.extend(f'   - {w}' for w in warnings)
        return log

    def _save_bom(self, product_name, code, lines, product_map, comp_map, key_fn):
        warnings = []

        tmpl = product_map.get(product_name)
        if tmpl is None:
            tmpl = self.env['product.template'].search([('name', '=', product_name)], limit=1)
        if not tmpl:
            warnings.append(f'Produit fini "{product_name}" introuvable, BOM ignorée')
            return 'error', warnings

        domain = [('product_tmpl_id', '=', tmpl.id)]
        if code:
            domain.append(('code', '=', code))
        existing_bom = self.env['mrp.bom'].search(domain, limit=1)
        if existing_bom:
            if existing_bom.bom_line_ids:
                warnings.append(
                    f'BOM "{product_name}" (réf: {code or "—"}) déjà existante avec composants, ignorée'
                )
                return 'skipped', warnings
            existing_bom.unlink()

        with self.env.cr.savepoint():
            bom_vals = {'product_tmpl_id': tmpl.id, 'type': 'normal'}
            if code:
                bom_vals['code'] = code
            bom = self.env['mrp.bom'].create(bom_vals)

            for line in lines:
                key = key_fn(line['reference'], line['nom'])
                variant = comp_map.get(key)
                if variant is None:
                    warnings.append(
                        f'Ligne {line["row"]} : composant "{line["nom"]}" absent du comp_map, ignoré'
                    )
                    continue
                self.env['mrp.bom.line'].create({
                    'bom_id': bom.id,
                    'product_id': variant.id,
                    'product_qty': line['qty'],
                })

        return 'created', warnings

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _map_product_type(self, type_str):
        # Odoo 19: type is 'consu' (Goods), 'service', or 'combo'
        # Old 'product'/'storable' values map to 'consu'
        mapping = {'consu': 'consu', 'service': 'service', 'combo': 'combo'}
        return mapping.get((type_str or '').strip(), 'consu')

    def _resolve_uom(self, uom_str):
        uom_str = (uom_str or '').strip()
        if uom_str == 'm':
            uom = self.env.ref('uom.product_uom_meter', raise_if_not_found=False)
            if uom:
                return uom
            uom = self.env['uom.uom'].search(
                ['|', ('name', '=', 'm'), ('name', 'ilike', 'mètre')], limit=1
            )
            if uom:
                return uom
        elif uom_str:
            uom = self.env['uom.uom'].search(
                ['|', ('name', 'ilike', 'unit'), ('name', 'ilike', 'unité')], limit=1
            )
            if uom:
                return uom
        return self.env.ref('uom.product_uom_unit')
