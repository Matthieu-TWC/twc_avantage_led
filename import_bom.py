#!/usr/bin/env python3
"""
Import de nomenclatures depuis Excel vers Odoo via XML-RPC.

Usage:
    export ODOO_URL=https://avantage-led-test-import.odoo.com
    export ODOO_DB=avantage-led-test-import
    export ODOO_LOGIN=you@example.com
    export ODOO_API_KEY=your_api_key

    python3 import_bom.py chemin/vers/Nomenclature.xlsx
"""

import os
import sys
import xmlrpc.client

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl manquant. Installez-le : pip3 install openpyxl")

URL     = os.environ.get('ODOO_URL', '').rstrip('/')
DB      = os.environ.get('ODOO_DB', '')
LOGIN   = os.environ.get('ODOO_LOGIN', '')
API_KEY = os.environ.get('ODOO_API_KEY', '')


def main():
    # ── Vérifications ─────────────────────────────────────────────────────
    missing = [k for k, v in {'ODOO_URL': URL, 'ODOO_DB': DB,
                               'ODOO_LOGIN': LOGIN, 'ODOO_API_KEY': API_KEY}.items() if not v]
    if missing:
        sys.exit(f"Variables d'environnement manquantes : {', '.join(missing)}")

    if len(sys.argv) < 2:
        sys.exit("Usage : python3 import_bom.py chemin/vers/fichier.xlsx")

    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        sys.exit(f"Fichier introuvable : {excel_path}")

    # ── Connexion XML-RPC ─────────────────────────────────────────────────
    print(f"→ Connexion à {URL} ...")
    common = xmlrpc.client.ServerProxy(f'{URL}/xmlrpc/2/common')
    uid = common.authenticate(DB, LOGIN, API_KEY, {})
    if not uid:
        sys.exit("Authentification échouée. Vérifiez URL, DB, login et clé API.")
    print("  Connecté ✔")

    models = xmlrpc.client.ServerProxy(f'{URL}/xmlrpc/2/object')

    def rpc(model, method, args=None, kw=None):
        return models.execute_kw(DB, uid, API_KEY, model, method, args or [], kw or {})

    # ── Helpers ───────────────────────────────────────────────────────────
    all_uoms = rpc('uom.uom', 'search_read', [[]], {'fields': ['id', 'name']})
    uom_by_name = {u['name'].lower(): u['id'] for u in all_uoms}
    unit_id = next(
        (u['id'] for u in all_uoms if 'unit' in u['name'].lower() or 'unité' in u['name'].lower()),
        all_uoms[0]['id'] if all_uoms else 1,
    )

    def resolve_uom(uom_str):
        s = (uom_str or '').strip().lower()
        if s in uom_by_name:
            return uom_by_name[s]
        for name, uid_ in uom_by_name.items():
            if s and (s in name or name in s):
                return uid_
        return unit_id

    def map_type(type_str):
        return {'consu': 'consu', 'service': 'service', 'combo': 'combo'}.get(
            (type_str or '').strip(), 'consu'
        )

    def bom_key(ref, nom):
        return f'{ref}|{nom}' if ref else nom

    def str_cell(v):
        return str(v).strip() if v not in (None, '') else ''

    def float_cell(v):
        try:
            return float(v) if v not in (None, '') else None
        except (ValueError, TypeError):
            return None

    wb = load_workbook(excel_path, data_only=True)
    log = []

    # ── Phase 0 : Partenaires ─────────────────────────────────────────────
    print("→ Import des partenaires ...")
    sheet = wb['Partenaires']
    existing = rpc('res.partner', 'search_read', [[]], {'fields': ['id', 'name']})
    partner_by_name = {p['name']: p['id'] for p in existing}
    partner_map = {}
    created = already = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        nom, est_entreprise, _ = (row + (None,) * 3)[:3]
        nom = str_cell(nom)
        if not nom:
            continue
        if nom in partner_by_name:
            partner_map[nom] = partner_by_name[nom]
            already += 1
        else:
            pid = rpc('res.partner', 'create', [{'name': nom, 'is_company': bool(est_entreprise)}])
            partner_by_name[nom] = pid
            partner_map[nom] = pid
            created += 1

    log.append(f'✔ Partenaires importés : {created} créés, {already} déjà existants')
    print(f"  {log[-1]}")

    # ── Phase 1 : Produits finis ──────────────────────────────────────────
    print("→ Import des produits finis ...")
    sheet = wb['Produits finis']
    existing = rpc('product.template', 'search_read', [[]], {'fields': ['id', 'name', 'default_code']})
    tmpl_by_code = {t['default_code']: t['id'] for t in existing if t['default_code']}
    tmpl_by_name = {t['name']: t['id'] for t in existing}
    tmpl_no_code = {t['id'] for t in existing if not t['default_code']}
    product_map = {}
    created = already = updated = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        nom, ref, uom_str, ptype = (row + (None,) * 4)[:4]
        nom = str_cell(nom)
        if not nom:
            continue
        ref = str_cell(ref)

        tmpl_id = tmpl_by_code.get(ref) if ref else None
        if tmpl_id is None:
            tmpl_id = tmpl_by_name.get(nom)

        if tmpl_id:
            if ref and tmpl_id in tmpl_no_code:
                rpc('product.template', 'write', [[tmpl_id], {'default_code': ref}])
                tmpl_by_code[ref] = tmpl_id
                tmpl_no_code.discard(tmpl_id)
                updated += 1
            already += 1
        else:
            vals = {'name': nom, 'type': map_type(ptype), 'uom_id': resolve_uom(uom_str)}
            if ref:
                vals['default_code'] = ref
            tmpl_id = rpc('product.template', 'create', [vals])
            tmpl_by_name[nom] = tmpl_id
            if ref:
                tmpl_by_code[ref] = tmpl_id
            created += 1

        product_map[f'[{ref}] {nom}' if ref else nom] = tmpl_id
        product_map[nom] = tmpl_id

    log.append(f'✔ Produits finis       : {created} créés, {already} déjà existants, {updated} référence(s) mise(s) à jour')
    print(f"  {log[-1]}")

    # ── Phase 2 : Composants ──────────────────────────────────────────────
    print("→ Import des composants ...")
    sheet = wb['Nomenclatures']

    # Reload templates — ils ont pu augmenter avec les produits finis
    existing = rpc('product.template', 'search_read', [[]], {'fields': ['id', 'name', 'default_code']})
    tmpl_by_code = {t['default_code']: t['id'] for t in existing if t['default_code']}
    tmpl_by_name = {t['name']: t['id'] for t in existing}
    tmpl_no_code = {t['id'] for t in existing if not t['default_code']}
    tmpl_id_to_code = {t['id']: t['default_code'] for t in existing if t['default_code']}

    comp_map = {}  # key -> variant_id
    created = already = updated = 0

    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not any(row):
            continue
        (_, _, comp_nom, comp_ref, comp_uom,
         _qty, comp_prix, comp_poids, comp_four, _) = (row + (None,) * 10)[:10]

        nom = str_cell(comp_nom)
        if not nom:
            continue
        ref = str_cell(comp_ref)
        k = bom_key(ref, nom)
        if k in comp_map:
            continue

        tmpl_id = tmpl_by_code.get(ref) if ref else None
        if tmpl_id is None:
            tmpl_id = tmpl_by_name.get(nom)

        if tmpl_id:
            if ref and tmpl_id in tmpl_no_code:
                rpc('product.template', 'write', [[tmpl_id], {'default_code': ref}])
                tmpl_by_code[ref] = tmpl_id
                tmpl_no_code.discard(tmpl_id)
                updated += 1
            already += 1
        else:
            vals = {'name': nom, 'type': 'consu', 'uom_id': resolve_uom(comp_uom)}
            if ref:
                vals['default_code'] = ref
            prix = float_cell(comp_prix)
            poids = float_cell(comp_poids)
            if prix is not None:
                vals['standard_price'] = prix
            if poids is not None:
                vals['weight'] = poids
            tmpl_id = rpc('product.template', 'create', [vals])
            tmpl_by_name[nom] = tmpl_id
            if ref:
                tmpl_by_code[ref] = tmpl_id
            created += 1

        variants = rpc('product.product', 'search_read',
                       [[['product_tmpl_id', '=', tmpl_id]]],
                       {'fields': ['id'], 'limit': 1})
        if not variants:
            print(f"  ⚠ Pas de variante pour '{nom}', ignoré")
            continue
        comp_map[k] = variants[0]['id']

        # Lier fournisseur
        fournisseur = str_cell(comp_four)
        if fournisseur and fournisseur in partner_map:
            partner_id = partner_map[fournisseur]
            existing_sellers = rpc('product.supplierinfo', 'search',
                                   [[['partner_id', '=', partner_id],
                                     ['product_tmpl_id', '=', tmpl_id]]])
            if not existing_sellers:
                prix = float_cell(comp_prix)
                rpc('product.template', 'write', [[tmpl_id], {'seller_ids': [(0, 0, {
                    'partner_id': partner_id,
                    'price': prix or 0.0,
                })]}])

    log.append(f'✔ Composants importés  : {created} créés, {already} déjà existants, {updated} référence(s) mise(s) à jour')
    print(f"  {log[-1]}")

    # ── Phase 3 : BOMs ────────────────────────────────────────────────────
    print("→ Import des nomenclatures ...")
    sheet = wb['Nomenclatures']
    warnings = []
    bom_created = bom_skipped = 0
    current_product = current_code = None
    current_lines = []

    def flush_bom():
        nonlocal bom_created, bom_skipped
        if current_product is None:
            return
        tmpl_id = product_map.get(current_product)
        if tmpl_id is None:
            warnings.append(f'Produit fini "{current_product}" introuvable, BOM ignorée')
            return

        # Use product's default_code as BOM code when none is specified in Excel
        effective_code = current_code or tmpl_id_to_code.get(tmpl_id, '')
        if not current_code and effective_code:
            print(f"  ℹ BOM '{current_product}' : code auto-défini à '{effective_code}' (default_code du produit)")

        domain = [['product_tmpl_id', '=', tmpl_id]]
        if effective_code:
            domain.append(['code', '=', effective_code])
        existing_boms = rpc('mrp.bom', 'search_read', [domain],
                            {'fields': ['id', 'bom_line_ids'], 'limit': 1})
        if existing_boms:
            if existing_boms[0]['bom_line_ids']:
                warnings.append(f'BOM "{current_product}" déjà existante avec composants, ignorée')
                bom_skipped += 1
                return
            rpc('mrp.bom', 'unlink', [[existing_boms[0]['id']]])

        bom_vals = {'product_tmpl_id': tmpl_id, 'type': 'normal'}
        if effective_code:
            bom_vals['code'] = effective_code
        bom_id = rpc('mrp.bom', 'create', [bom_vals])

        for line in current_lines:
            k = bom_key(line['ref'], line['nom'])
            variant_id = comp_map.get(k)
            if variant_id is None:
                warnings.append(f'Composant "{line["nom"]}" absent, ligne ignorée')
                continue
            rpc('mrp.bom.line', 'create', [{'bom_id': bom_id,
                                             'product_id': variant_id,
                                             'product_qty': line['qty']}])
        bom_created += 1

    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not any(row):
            continue
        (bom_pf, bom_ref, comp_nom, comp_ref, _,
         comp_qty, _, _, _, _) = (row + (None,) * 10)[:10]

        bom_pf = str_cell(bom_pf)
        comp_nom = str_cell(comp_nom)
        comp_ref = str_cell(comp_ref)

        if bom_pf:
            flush_bom()
            current_product = bom_pf
            current_code = str_cell(bom_ref)
            current_lines = []

        if comp_nom:
            qty = float_cell(comp_qty) or 1.0
            current_lines.append({'nom': comp_nom, 'ref': comp_ref, 'qty': qty})

    flush_bom()

    log.append(f'✔ BOMs importées       : {bom_created} créées, {bom_skipped} déjà existantes')
    if warnings:
        log.append(f'⚠ Avertissements ({len(warnings)}) :')
        log.extend(f'   - {w}' for w in warnings)

    print(f"  {log[-1] if not warnings else log[-2]}")

    # ── Phase 4 : Correction des BOMs existantes sans référence ───────────
    print("→ Correction des BOMs existantes sans référence ...")
    boms_no_code = rpc('mrp.bom', 'search_read',
                       [[['code', '=', False]]],
                       {'fields': ['id', 'product_tmpl_id']})
    bom_fixed = 0
    for bom in boms_no_code:
        if not bom['product_tmpl_id']:
            continue
        tmpl_id = bom['product_tmpl_id'][0]
        tmpl_name = bom['product_tmpl_id'][1]
        default_code = tmpl_id_to_code.get(tmpl_id)
        if not default_code:
            tmpl_data = rpc('product.template', 'read', [[tmpl_id]], {'fields': ['default_code']})
            default_code = tmpl_data[0]['default_code'] if tmpl_data else None
        if default_code:
            rpc('mrp.bom', 'write', [[bom['id']], {'code': default_code}])
            print(f"  ✏ BOM #{bom['id']} ({tmpl_name}) : référence mise à jour → '{default_code}'")
            bom_fixed += 1

    log.append(f'✔ BOMs sans référence  : {bom_fixed} corrigée(s)')
    print(f"  {log[-1]}")

    print()
    print('── Rapport final ────────────────────────')
    for line in log:
        print(line)


if __name__ == '__main__':
    main()
